from openpyxl import load_workbook
from openpyxl.styles import Alignment

excel_path = "cat_MLM4620_stats.xlsx"
wb = load_workbook(excel_path)
ws = wb.active

# 插入第一行
ws.insert_rows(1)

# 体量及特性
ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=11)
top_left = ws.cell(row=1, column=3)
top_left.value = "体量及特性"
top_left.alignment = Alignment(horizontal="center", vertical="center")

# 竞争性-品牌&店铺垄断性
ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=13)
top_left = ws.cell(row=1, column=12)
top_left.value = "竞争性-品牌&店铺垄断性"
top_left.alignment = Alignment(horizontal="center", vertical="center")

# 竞争性-FULL仓
ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=17)
top_left = ws.cell(row=1, column=14)
top_left.value = "竞争性-FULL仓"
top_left.alignment = Alignment(horizontal="center", vertical="center")

# 新品机会
ws.merge_cells(start_row=1, start_column=18, end_row=1, end_column=26)
top_left = ws.cell(row=1, column=18)
top_left.value = "新品机会"
top_left.alignment = Alignment(horizontal="center", vertical="center")

# 广告及其他
ws.merge_cells(start_row=1, start_column=27, end_row=1, end_column=28)
top_left = ws.cell(row=1, column=27)
top_left.value = "广告及其他"
top_left.alignment = Alignment(horizontal="center", vertical="center")

# 保存最终 Excel
wb.save("cat_MLM4620_stats_merged.xlsx")
print("导出完成：cat_MLM4620_stats_merged.xlsx")
