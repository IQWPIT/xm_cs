import os
from datetime import datetime, timedelta
from calendar import monthrange
from pymongo import MongoClient
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
from collections import defaultdict
from tqdm import tqdm

os.environ["NET"] = "TUNNEL"
os.environ["NET3"] = "NXQ"

from dm.connector.mongo.manager3 import get_collection

# ================= Mongo 连接 =================
visual_plus = get_collection("main_ml_mx", "ml_mx", "visual_plus")
sku_col = get_collection("main_ml_mx", "ml_mx", "sku")
every_day_sku = get_collection("main_ml_mx", "ml_mx", "every_day_sku")
visualize_table = get_collection("main_ml_mx", "ml_mx", "visualize_table")

uri = "mongodb://common:ase5yDFHG%24%25FDSdif%40%23GH@localhost:37031/admin?authMechanism=SCRAM-SHA-1&directConnection=true&readPreference=primary"
client = MongoClient(uri)
scrapy_buffer = client["ml_scrapy_buffer"]["ml_mx"]

mongo_uri = "mongodb://erp:Damai20230214*.*@42.193.215.253:27017/?authMechanism=DEFAULT&authSource=erp&directConnection=true"
client = MongoClient(mongo_uri)
db = client["erp"]
bi_site_data = db["bi_category_data"]

# ================= 参数 =================
cat_ids = [
     "MLM4620","MLM172387",
    "MLM178702"]
cat_ids = visual_plus.find_one({"cat_id":"MLM436380"},{"_id":0,"cat":1})["cat"]
time_list = [
    202411,202412,202501,202502,202503,202504,202505,202506,202507,202508,202509,
    202510,
    202511]

# ================= 工具函数 =================
def get_nested(d, key, default=0):
    cur = d
    for k in key.split("."):
        if isinstance(cur, dict) and k in cur:
            cur = cur[k]
        else:
            return default
    return cur

def to_round(v):
    try:
        return round(float(v), 2)
    except Exception:
        return 0.00

def to_percent_str(v):
    try:
        return f"{round(float(v) * 100, 2)}%"
    except Exception:
        return "0.00%"
def parse_puton_date(v):
    if isinstance(v, datetime):
        return v
    try:
        v = int(v)
        if 20200101 <= v <= 20991231:
            return datetime.strptime(str(v), "%Y%m%d")
    except:
        return None
    return None

def pct(val):
    return f"{round(val*100,2)}%" if val else "0%"

def compute_sales_growth(sku_id, curr_month, sku_sales_history, time_list, N=3):
    history = sku_sales_history.get(sku_id,{})
    prev_months = [m for m in sorted(time_list) if m < curr_month][-N:]
    prev_sales = [history.get(m,0) for m in prev_months if history.get(m,0) > 0]
    curr_sales = history.get(curr_month,0)
    if not prev_sales or curr_sales <= 0:
        return None
    avg_prev = sum(prev_sales)/len(prev_sales)
    return round(curr_sales/avg_prev,2) if avg_prev else None

# ================= dT 类型检测 =================
sample_dT = every_day_sku.find_one({}, {"dT":1,"_id":0})
dT_type = type(sample_dT["dT"]) if sample_dT and "dT" in sample_dT else None

# ================= 结果列表 =================
sku_sales_history = defaultdict(dict)

# ================= 主循环: 按月份 =================
all_results = []

for t in tqdm(time_list, desc="月份进度"):
    month_results = []
    for cat_id in tqdm(cat_ids, desc=f"{t} 月份 cat_id 进度", leave=False):
        data = visual_plus.find_one({"cat_id": cat_id}, {"top100_skus":1,"_id":0})
        if not data or "top100_skus" not in data:
            continue

        month_data = data["top100_skus"].get(str(t))
        if not month_data:
            continue

        sku_ids = list(month_data.keys())[:50]  # 每个 cat_id 取前50 SKU
        year, month = int(str(t)[:4]), int(str(t)[4:])
        month_start = datetime(year, month, 1)
        month_end = datetime(year, month, monthrange(year, month)[1])

        stats = defaultdict(float)
        stats.update({
            "month_sales":0,"month_gmv":0,"sku_cnt":0,
            "full_sales":0,"full_sku_cnt":0,
            "shop_set":set(),"brand_set":set(),
            "new_30_cnt":0,"new_30_sales":0,
            "new_90_cnt":0,"new_90_sales":0,
            "new_180_cnt":0,"new_180_sales":0,
            "rating_sum":0,"rating_cnt":0,
            "offers_sku_cnt":0,"offers_sales":0,
            "conversion_all_order":0,"conversion_all_view":0,
            "volume_m3_sum":0,"volume_m3_cnt":0,
            "volume_weight_sum":0,"volume_weight_cnt":0
        })

        # 批量查询 SKU 信息
        sku_infos = {doc["sku_id"]:doc for doc in sku_col.find(
            {"sku_id":{"$in":sku_ids}},
            {"sku_id":1,"active_price":1,"stock_type":1,"sellerName":1,"brand":1,
             "puton_date":1,"offersInf":1,"conversion_all_order":1,"conversion_all_view":1,
             "attributes":1,"_id":0}
        ) if "sku_id" in doc}

        # 批量查询评分
        rating_docs = {doc["sku_id"]:doc for doc in scrapy_buffer.find(
            {"sku_id":{"$in":sku_ids},"t":{"$gte":month_start,"$lte":month_end}},
            {"sku_id":1,"rating_avg":1,"_id":0}
        ) if "sku_id" in doc}

        for sku_id in sku_ids:
            sku_info = sku_infos.get(sku_id)
            if not sku_info:
                continue

            order = int(month_data.get(sku_id,{}).get("order",0))
            price = int(sku_info.get("active_price") or 0)
            gmv = order*price
            stock_type = sku_info.get("stock_type")
            seller = sku_info.get("sellerName")
            brand = sku_info.get("brand")
            puton_date = parse_puton_date(sku_info.get("puton_date"))

            stats["month_sales"] += order
            stats["month_gmv"] += gmv
            stats["sku_cnt"] += 1
            stats["conversion_all_order"] += int(sku_info.get("conversion_all_order") or 0)
            stats["conversion_all_view"] += int(sku_info.get("conversion_all_view") or 0)
            if seller: stats["shop_set"].add(seller)
            if brand: stats["brand_set"].add(brand)
            if stock_type=="ful":
                stats["full_sales"] += order
                stats["full_sku_cnt"] += 1

            # 新品统计
            if puton_date:
                NEW_WINDOWS = [(30,"new_30_cnt","new_30_sales"),
                               (90,"new_90_cnt","new_90_sales"),
                               (180,"new_180_cnt","new_180_sales")]
                for days,cnt_key,sales_key in NEW_WINDOWS:
                    new_start,new_end = puton_date, puton_date+timedelta(days=days)
                    overlap_start = max(new_start, month_start)
                    overlap_end = min(new_end, month_end+timedelta(days=1))
                    if overlap_start>=overlap_end:
                        continue

                    query = {"sl":sku_id}
                    if dT_type is datetime:
                        query["dT"]={"$gte":overlap_start,"$lt":overlap_end}
                    elif dT_type is int:
                        query["dT"]={"$gte":int(overlap_start.strftime("%Y%m%d")),
                                     "$lte":int((overlap_end-timedelta(days=1)).strftime("%Y%m%d"))}
                    elif dT_type is str:
                        query["dT"]={"$gte":overlap_start.strftime("%Y-%m-%d"),
                                     "$lte":(overlap_end-timedelta(days=1)).strftime("%Y-%m-%d")}
                    else:
                        continue

                    overlap_sales = next(every_day_sku.aggregate([
                        {"$match":query},
                        {"$group":{"_id":None,"total":{"$sum":"$oD"}}}
                    ]), {"total":0})["total"]

                    if overlap_sales>0:
                        stats[cnt_key] += 1
                        stats[sales_key] += overlap_sales

            # rating
            rating_doc = rating_docs.get(sku_id)
            if rating_doc and "rating_avg" in rating_doc:
                stats["rating_sum"] += float(rating_doc["rating_avg"])
                stats["rating_cnt"] += 1

            # offers
            offers = sku_info.get("offersInf")
            if offers and isinstance(offers,list) and offers:
                stats["offers_sku_cnt"] += 1
                stats["offers_sales"] += order

            # 体积&重量
            length=width=height=volume_weight_kg=None
            for attr in sku_info.get("attributes",[]):
                aid = attr.get("id")
                val = attr.get("v_name_en")
                if not aid or not val: continue
                nums = re.findall(r"[\d.]+", val)
                if not nums: continue
                try:
                    num=float(nums[0])
                except: continue
                val_lower = val.lower()
                if aid=="PACKAGE_LENGTH": num*=0.001 if "mm" in val_lower else 0.01 if "cm" in val_lower else 1; length=num
                elif aid=="PACKAGE_WIDTH": num*=0.001 if "mm" in val_lower else 0.01 if "cm" in val_lower else 1; width=num
                elif aid=="PACKAGE_HEIGHT": num*=0.001 if "mm" in val_lower else 0.01 if "cm" in val_lower else 1; height=num
                elif aid=="PACKAGE_WEIGHT": volume_weight_kg=num/1000 if "g" in val_lower else num

            if length and width and height:
                volume_m3 = length*width*height
                if volume_m3>0: stats["volume_m3_sum"]+=volume_m3; stats["volume_m3_cnt"]+=1
                if volume_weight_kg: stats["volume_weight_sum"]+=volume_weight_kg; stats["volume_weight_cnt"]+=1

            sku_sales_history[sku_id][t]=order

        # 销量增长指数 & 平均评分
        growth_indices = [compute_sales_growth(sku_id,t,sku_sales_history,time_list,3) for sku_id in sku_ids]
        growth_indices = [g for g in growth_indices if g is not None]
        sales_growth_index = round(sum(growth_indices)/len(growth_indices),2) if growth_indices else 0
        avg_rating = round(stats["rating_sum"]/stats["rating_cnt"],2) if stats["rating_cnt"] else 0

        #   全量
        collection_name = f"visualize_table_bak_{t}28"
        collection = get_collection("main_ml_mx", "ml_mx", collection_name)
        doc = collection.find_one({"cat_id": cat_id}) or {}
        gmv_doc = visualize_table.find_one({"cat_id": cat_id}, {"monthly_gmv_trend": 1, "_id": 0}) or {}
        gmv = gmv_doc.get("monthly_gmv_trend", {})

        p = visual_plus.find_one({"cat_id": cat_id}, {"offersInf": 1, "_id": 0}) or {}
        offers_inf = p.get("offersInf", {}).get(str(t), {})

        bi_doc = bi_site_data.find_one({"cat_id": cat_id, "month": str(t)}) or {}
        ctr = bi_doc.get("ctr", 0)
        cpc = bi_doc.get("cpc", 0)
        acos = bi_doc.get("acos", 0)
        refund_rate = bi_doc.get("refund_rate", 0)

        follow_order = get_nested(offers_inf, "follow.order", 0)
        non_follow_order = get_nested(offers_inf, "non_follow.order", 0)
        sum_order = follow_order + non_follow_order
        follow_p = follow_order / sum_order if sum_order else 0

        total_prod_num = doc.get("total_prod_num") or 1
        t_q = t
        time = str(t)[4::]
        month_results.append({
            "cat_id":cat_id,
            "month":t_q,
            #           体量 3-18
            f"{time}月_全量_销量": to_round(get_nested(doc, "sale_num.order30d")),
            f"{time}月_全量_销售额(比索)": to_round(gmv.get(str(t), 0)),
            f"{time}月_全量_成交平均价": to_round(get_nested(doc, "avg_price")),
            f"{time}月_全量_商品平均销量": to_round(get_nested(doc, "sale_num.order30d") / total_prod_num),
            f"{time}月_全量_销量增长指数": to_round(get_nested(doc, "increase_relative_ratio")),

            f"{time}月_top50_月销量":stats["month_sales"],
            f"{time}月_top50_月销售额(比索)":stats["month_gmv"],
            f"{time}月_top50_成交平均价":round(stats["month_gmv"]/stats["month_sales"],2) if stats["month_sales"] else 0,
            f"{time}月_top50_商品平均销量":round(stats["month_sales"]/stats["sku_cnt"],2) if stats["sku_cnt"] else 0,
            f"{time}月_top50_销量增长指数":pct(sales_growth_index),
            f"{time}月_top50_FULL仓销量占比":pct(stats["full_sales"]/stats["month_sales"] if stats["month_sales"] else 0),
            f"{time}月_top50_平均体积(m³)":round(stats["volume_m3_sum"]/stats["volume_m3_cnt"],6) if stats["volume_m3_cnt"] else 0,
            f"{time}月_top50_平均体积重量(kg)":round(stats["volume_weight_sum"]/stats["volume_weight_cnt"],3) if stats["volume_weight_cnt"] else 0,
            f"{time}月_top50_跟卖商品数量":stats["offers_sku_cnt"],
            f"{time}月_top50_跟卖商品销量":stats["offers_sales"],
            f"{time}月_top50_跟卖商品销量占比":pct(stats["offers_sales"]/stats["month_sales"] if stats["month_sales"] else 0),
            #           竞争性-品牌&店铺垄断性 19-27
            f"{time}月_全量_活跃商品占比": to_percent_str(get_nested(doc, "index_active")),
            f"{time}月_全量_跟卖销量占比": to_percent_str(follow_p),
            f"{time}月_全量_产品集中度": to_percent_str(get_nested(doc, "product_centralize_ratio.ratio")),
            f"{time}月_全量_品牌集中度": to_percent_str(get_nested(doc, "market_centralize_ratio.ratio")),
            f"{time}月_全量_店铺集中度": to_percent_str(get_nested(doc, "seller_centralize_data.ratio")),
            f"{time}月_全量_店铺数": to_round(get_nested(doc, "seller_info.seller_num")),
            f"{time}月_全量_品牌数": to_round(get_nested(doc, "brand_num")),

            f"{time}月_top50_店铺数":len(stats["shop_set"]),
            f"{time}月_top50_品牌数":len(stats["brand_set"]),

            #           FULL    28-38
            f"{time}月_全量_FULL仓销量占比": to_percent_str(get_nested(doc, "stock_info.fbm_order30d_ratio")),
            f"{time}月_全量_FULL仓商品数": to_round(get_nested(doc, "stock_info.fbm_number")),
            f"{time}月_全量_FULL仓商品数量占比": to_percent_str(get_nested(doc, "stock_info.fbm_ratio")),
            f"{time}月_全量_FULL仓活跃商品数": to_round(get_nested(doc, "stock_info.fbm_count30d")),
            f"{time}月_全量_FULL仓动销率": to_percent_str(get_nested(doc, "stock_info.full_stock_turnover_rate")),
            f"{time}月_全量_FULL仓近30天销量": to_round(get_nested(doc, "stock_info.fbm_order30d")),
            f"{time}月_全量_FULL仓动销平均销量": to_round(get_nested(doc, "stock_info.full_avg_sales")),

            f"{time}月_top50_FULL商品数":stats["full_sku_cnt"],
            f"{time}月_top50_FULL仓商品数量占比":pct(stats["full_sku_cnt"]/stats["sku_cnt"] if stats["sku_cnt"] else 0),
            f"{time}月_top50_FULL仓近30天销量":stats["full_sales"],
            f"{time}月_top50_FULL仓平均销量":round(stats["full_sales"]/stats["full_sku_cnt"],2) if stats["full_sku_cnt"] else 0,
            #           新品  39-56
            f"{time}月_全量_近30天新品数量": to_round(get_nested(doc, "new_product_info.p30d_count30d")),
            f"{time}月_全量_近30天新品销量": to_round(get_nested(doc, "new_product_info.p30d_order30d")),
            f"{time}月_全量_近30天新品销量占比": to_percent_str(get_nested(doc, "new_product_info.p30d_order30d_ratio")),
            f"{time}月_全量_近90天新品数量": to_round(get_nested(doc, "new_product_info.p90d_count30d")),
            f"{time}月_全量_近90天新品销量": to_round(get_nested(doc, "new_product_info.p90d_order30d")),
            f"{time}月_全量_近90天新品销量占比": to_percent_str(get_nested(doc, "new_product_info.p90d_order30d_ratio")),
            f"{time}月_全量_近180天新品数量": to_round(get_nested(doc, "new_product_info.p180d_count30d")),
            f"{time}月_全量_近180天新品销量": to_round(get_nested(doc, "new_product_info.p180d_order30d")),
            f"{time}月_全量_近180天新品销量占比": to_percent_str(get_nested(doc, "new_product_info.p180d_order30d_ratio")),

            f"{time}月_top50_近30天新品数量":stats["new_30_cnt"],
            f"{time}月_top50_近30天新品销量":stats["new_30_sales"],
            f"{time}月_top50_近30天新品销量占比":pct(stats["new_30_sales"]/stats["month_sales"] if stats["month_sales"] else 0),
            f"{time}月_top50_近90天新品数量":stats["new_90_cnt"],
            f"{time}月_top50_近90天新品销量":stats["new_90_sales"],
            f"{time}月_top50_近90天新品销量占比":pct(stats["new_90_sales"]/stats["month_sales"] if stats["month_sales"] else 0),
            f"{time}月_top50_近180天新品数量":stats["new_180_cnt"],
            f"{time}月_top50_近180天新品销量":stats["new_180_sales"],
            f"{time}月_top50_近180天新品销量占比":pct(stats["new_180_sales"]/stats["month_sales"] if stats["month_sales"] else 0),
            #           广告  57-62
            f"{time}月_全量_广告点击率": to_percent_str(ctr),
            f"{time}月_全量_平均单词点击广告费(美元)": to_round(cpc),
            f"{time}月_全量_平均广告销售成本比": acos,
            f"{time}月_全量_平均退款率": to_percent_str(refund_rate),

            f"{time}月_top50_平均评分":avg_rating,
            f"{time}月_top50_平均转换率":pct(stats["conversion_all_order"]/stats["conversion_all_view"] if stats["conversion_all_view"] else 0),

        })

    all_results.extend(month_results)

# ================= 导出 Excel，每个月一个 sheet =================
df = pd.DataFrame(all_results)
excel_path="全量_top50.xlsx"
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    for month in time_list:
        df_month = df[df["month"]==month]
        if not df_month.empty:
            df_month.to_excel(writer,index=False,sheet_name=str(month))

# ================= 合并表头 =================
wb = load_workbook(excel_path)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=18)
    ws.cell(1, 3, "体量及特性").alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=19, end_row=1, end_column=27)
    ws.cell(1, 19, "竞争性-品牌&店铺垄断性").alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=28, end_row=1, end_column=38)
    ws.cell(1, 28, "竞争性-FULL仓").alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=39, end_row=1, end_column=56)
    ws.cell(1, 39, "新品机会").alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=57, end_row=1, end_column=62)
    ws.cell(1, 57, "广告及其他").alignment = Alignment(horizontal="center", vertical="center")

wb.save(excel_path)
print(f"导出完成，每个月份一个 sheet，多个 cat_id 数据在同一表：{excel_path}")
